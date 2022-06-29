from  openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class writexlsx:

    def __init__(self,excel_name) -> None:
        wb = Workbook()
        ws1 = wb.create_sheet("info")
        ws1.sheet_properties.tabColor = "1072BA"

        # insert date
        ws1['A1'] = 'Date:'
        datetime_dt = datetime.today()# 獲得當地時間
        datetime_str = datetime_dt.strftime("%Y/%m/%d %H:%M:%S")
        ws1['B1'] = datetime_str
        
        self.name = './'+str(excel_name)+'.xlsx'
        wb.save(self.name)


    def add_info(self,pt_a_path,pt_b_path,a_starting_frame,b_starting_frame,common_starting_frame,a_total,b_total,common_total):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['info']
        ws['A2']='data_set_a:'
        ws['A3']='data_set_b:'
        ws['A4']='pt_a_starting_frame:'
        ws['A5']='pt_b_starting_frame:'
        ws['A6']='common_starting_frame:'
        ws['A7']='pt_a_total_frame:'
        ws['A8']='pt_b_total_frame:'
        ws['A9']='common_total_frame:'

        ws['B2']=pt_a_path
        ws['B3']=pt_b_path
        ws['B4']=a_starting_frame
        ws['B5']=b_starting_frame
        ws['B6']=common_starting_frame
        ws['B7']=a_total
        ws['B8']=b_total
        ws['B9']=common_total

        self.a_start= int(a_starting_frame)
        self.b_start= int(b_starting_frame)
        self.common_frame = int(common_starting_frame)
        self.common_frame_stop = int(common_starting_frame)+int(common_total)
        wb.save(self.name)

    def add_data(self,list_a,list_b): 
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb.create_sheet('data')
        
        ws['A1'] = 'global_frame'
        ws['B1'] = 'pt_a_x'
        ws['C1'] = 'pt_b_x'
        ws['D1'] = 'pt_a_y'
        ws['E1'] = 'pt_b_y'
        ws['F1'] = 'distance'
         

        # frame number
        for k in range (0,3000):
            str_frame = str(k)
            pos_col_a = 'A'+str(k+2)
            ws[pos_col_a] = str_frame

        for i in range (0,len(list_a)):
            pos_col_b = 'B'+str(i+2+self.a_start)
            pos_col_d = 'D'+str(i+2+self.a_start)
            ws[pos_col_b] = list_a[i][0]
            ws[pos_col_d] = list_a[i][1]

        for j in range (0,len(list_b)):
            pos_col_c = 'C'+str(j+2+self.b_start)
            pos_col_e = 'E'+str(j+2+self.b_start)
            ws[pos_col_c] = list_b[j][0]
            ws[pos_col_e] = list_b[j][1]

        # for rows in list_a:
        #     ws.append(rows)       
        wb.save(self.name)


    def add_distance(self,list_distance):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        
        for i in range (0,len(list_distance)):
            pos_col_f = 'F'+str(i+2+self.common_frame)
            ws[pos_col_f] = list_distance[i]
        wb.save(self.name)


    def add_plot_graph(self):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        chart = ScatterChart()
        chart.title = "Change of Distance"
        chart.style = 10
        chart.x_axis.title = 'Frame'
        chart.y_axis.title = 'Distance (Pixel)'

        xvalues = Reference(ws, min_col=1, min_row=2, max_row=int(self.common_frame_stop-1+2)) # range of x value

        values = Reference(ws, min_col=6, min_row=1, max_row=int(self.common_frame_stop-1+2)) # range of y value
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

        # 第一條散點
        s1 = chart.series[0]
        # 散點標記型別  'auto', 'dash', 'triangle', 'square', 'picture', 'circle', 'dot', 'plus', 'star', 'diamond', 'x'
        s1.marker.symbol = "circle"
        s1.marker.graphicalProperties.solidFill = "0000FF"  # Marker filling 設定標記填充的顏色
        s1.marker.graphicalProperties.line.solidFill = "0000FF"  # Marker outline 標記輪廓的顏色
        s1.graphicalProperties.line.noFill = True  # 關閉連線填充
        s1.graphicalProperties

        ws.add_chart(chart, "A10")
        wb.save(self.name)


# def edit():
#     file_home = './demo33.xlsx'
#     wb = load_workbook(filename=file_home)
#     ws = wb.create_sheet('created_sheet')
#     data = [[1,2,3],[4,5,6],[7,8,9,]]
#     ws.cell(row=4,column=1)
#     for rows in data:
#         ws.append(rows)
#     wb.save('./demo33.xlsx')

#======================================================
# job = writexlsx('3_0_th_7_0_th')
# job.add_info('./after_first_filter/feature_point_0/3_0_th.txt','./after_first_filter/feature_point_0/7_0_th.txt',3,5,5,4,4,2)
# job.add_data([[10,10],[10,10],[10,10],[10,10]],[[11,11],[11,11],[11,11],[11,11]])
# job.add_distance([1,2])
# job.add_plot_graph()