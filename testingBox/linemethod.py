import glob
import re
from itertools import combinations
import math
import cv2
import os


from openpyxl.drawing.image import Image

from  openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

# given --> common starting frame = 180 , name(aka path of excel)
# +2
# total = csf+tf-1-1
class testing:
    def __init__(self,name) -> None:

        self.name = name

    def test1(self):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        
        for i in range (180+2,180+2150-2-1):
            target_box_0 = 'F'+str(i)
            target_box_1 = 'F'+str(i+1)
            target_box_2 = 'F'+str(i+2)
            target_box_3 = 'F'+str(i+3)
            target_box_4 = 'F'+str(i+4)
            list_target_box_temp=[ws[target_box_0].value,ws[target_box_1].value,ws[target_box_2].value,ws[target_box_3].value,ws[target_box_4].value]
            target_min=list_target_box_temp.index(min(list_target_box_temp))
            print(i+target_min)
            group_in_5 = 'G'+str(i)
            ws[group_in_5] = int(i+target_min-2)
            # print(target_box)
            # print(ws[target_box].value)  #---> float(when single box)
        wb.save(self.name)

    def test2(self):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        for i in range (180+2,180+2150-2-1-4):
            target_box_0 = 'G'+str(i)
            target_box_1 = 'G'+str(i+1)
            target_box_2 = 'G'+str(i+2)
            target_box_3 = 'G'+str(i+3)
            target_box_4 = 'G'+str(i+4)
            group_in_5 = 'H'+str(i)
            if target_box_1>target_box_0 and target_box_2>target_box_0 and target_box_3>target_box_0 and target_box_4>target_box_0:
                ws[group_in_5] = int(0)
            else:
                ws[group_in_5] = int(1)
        
        wb.save(self.name)


    def test3(self):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        
        for i in range (180+2,180+2150-2-1):
            target_box_0 = 'G'+str(i)
            target_box_1 = 'G'+str(i+1)
            target_box_2 = 'G'+str(i+2)
            target_box_3 = 'G'+str(i+3)
            target_box_4 = 'G'+str(i+4)
            # print(ws[target_box_0].value)
            # print(type(ws[target_box_0].value))
            list_target_box_temp=[ws[target_box_0].value,ws[target_box_1].value,ws[target_box_2].value,ws[target_box_3].value,ws[target_box_4].value]
            target_min=list_target_box_temp.index(min(list_target_box_temp))
            print(i+target_min)
            group_in_5 = 'G'+str(i)
            ws[group_in_5] = int(i+target_min-2)
            # print(target_box)
            # print(ws[target_box].value)  #---> float(when single box)
        wb.save(self.name)




# testing('./testingBox/pt_diff_0_1_th_16_1_th.xlsx').test3()


