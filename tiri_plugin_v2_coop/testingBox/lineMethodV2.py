import glob
import re
from itertools import combinations
import math
import cv2
import os


from openpyxl.drawing.image import Image

from  openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)



#==============================================================================================================


class lineMethodV2:
    def __init__(self,switch) -> None:

        self.list_global = []
        self.list_pts = []
        self.list_pts_comb = []
        self.list_pts_comb1_2 =[]
        self.list_pts_comb3_2 = []
        self.list_pts_comb4_2 = []
        self.switch_lineMethod = int(switch)

        if switch == 0:
            self.path_global_frame = './feature_point_0/global_frame_number_0.txt'
            self.path_pts = './after_first_filter/feature_point_0/'
            self.path_glob_pts = './after_first_filter/feature_point_0/*.txt'
            self.path_glob_pts_base = './after_first_filter/feature_point_0/region_2/*.txt'
            self.path_glob_pts_1 = './after_first_filter/feature_point_0/region_1/*.txt'
            self.path_glob_pts_3 = './after_first_filter/feature_point_0/region_3/*.txt'
            self.path_glob_pts_4 = './after_first_filter/feature_point_0/region_4/*.txt'
        if switch == 1:
            self.path_global_frame = './feature_point_1/global_frame_number_1.txt'
            self.path_pts = './after_first_filter/feature_point_1/'
            self.path_glob_pts = './after_first_filter/feature_point_1/*.txt'
            self.path_glob_pts_base = './after_first_filter/feature_point_1/region_2/*.txt'
            self.path_glob_pts_1 = './after_first_filter/feature_point_1/region_1/*.txt'
            self.path_glob_pts_3 = './after_first_filter/feature_point_1/region_3/*.txt'
            self.path_glob_pts_4 = './after_first_filter/feature_point_1/region_4/*.txt'


    def run(self):

        terms_job = 0
        if self.switch_lineMethod ==0:
            folder_after_folder = './after_first_filter/feature_point_0/*'
        elif self.switch_lineMethod ==1:
            folder_after_folder = './after_first_filter/feature_point_1/*'

        for  comb in self.list_pts_comb: #

            comb_0,comb_1 = comb
            comb_0 = comb_0.replace('\\','/')
            comb_1 = comb_1.replace('\\','/')
            
            print('=========COMB0:',comb_0)
            print('=========COMB1:',comb_1)
            terms_job+=1
            print(str(terms_job)+'/'+str((math.comb(len(glob.glob(folder_after_folder)),2)))) # <-- progress bar

            list_pt_a = []
            list_pt_b = []

            with open(comb_0) as f1:
                for line in f1.readlines():
                    # print(line)    # testing
                    # number = [int(temp)for temp in line.split() if temp.isdigit()]
                    coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                    list_pt_a.append(coor) # it contains whole coor of point A

            with open(comb_1) as f2:
                for line2 in f2.readlines():
                    # print(line)    # testing
                    # number = [int(temp)for temp in line.split() if temp.isdigit()]
                    coor2 = re.findall("\d+\.\d+", line2)    # format:string    [convect from string with dot to float]
                    list_pt_b.append(coor2) # it contains whole coor of point B

            for i in range (len(self.list_global)):
                if float(list_pt_a[0][0]) == float(self.list_global[i][0]) and float(list_pt_a[0][1]) == float(self.list_global[i][1]):
                    print('x,y,frame_a: ',self.list_global[i]) # testing
                    pt_a_starting_frame = int(self.list_global[i][2]) # getting the starting frame number of pt A
                    print('pt_a_starting_frame:',pt_a_starting_frame)
                    break
            
            for j in range (len(self.list_global)):
                if float(list_pt_b[0][0]) == float(self.list_global[j][0]) and float(list_pt_b[0][1]) == float(self.list_global[j][1]):
                    print('x,y,frame_b: ',self.list_global[j]) # testing
                    pt_b_starting_frame = int(self.list_global[j][2]) # getting the starting frame number of pt B
                    print('pt_b_starting_frame:',pt_b_starting_frame)
                    break
            
            # getting common starting pt
            if pt_a_starting_frame > pt_b_starting_frame:
                common_starting_pt = pt_a_starting_frame
                flag_starting_frame = 0 # pt_a happens later than pt_b
            elif pt_a_starting_frame < pt_b_starting_frame:
                common_starting_pt = pt_b_starting_frame
                flag_starting_frame = 1 # pt_b happens later than pt_a
            elif pt_a_starting_frame == pt_b_starting_frame:
                common_starting_pt = pt_a_starting_frame
                flag_starting_frame = 2 # happen at the same time
            print('common starting point:',common_starting_pt)
            print('flag starting point:',flag_starting_frame)

            # getting common end pt
            if flag_starting_frame == 0:
                ans=(int(len(list_pt_b))-int(common_starting_pt-pt_b_starting_frame))-int(len(list_pt_a))
                common_diff = int(common_starting_pt-pt_b_starting_frame)
                if ans < 0:
                    total_frame = int(len(list_pt_b))-int(common_starting_pt-pt_b_starting_frame)
                elif ans > 0:
                    total_frame = int(len(list_pt_a))
                elif ans == 0:
                    total_frame = int(len(list_pt_a))

            if flag_starting_frame == 1:
                ans=(int(len(list_pt_a))-(int(common_starting_pt)-int(pt_a_starting_frame)))-int(len(list_pt_b))
                common_diff = int(common_starting_pt)-int(pt_a_starting_frame)
                if ans < 0:
                    total_frame = int(len(list_pt_a))-(int(common_starting_pt)-int(pt_a_starting_frame))
                elif ans > 0:
                    total_frame = int(len(list_pt_b))
                elif ans == 0:
                    total_frame = int(len(list_pt_b))
            
            if flag_starting_frame == 2:
                ans=int(len(list_pt_a))-int(len(list_pt_b))
                common_diff = 0
                if ans >0:
                    total_frame = int(len(list_pt_b))
                elif ans <0:
                    total_frame = int(len(list_pt_a))
                elif ans == 0:
                    total_frame = int(len(list_pt_a))
            
            print('total frame:',total_frame)

            # naming plugin
            comb_0_a_naming = comb_0.replace(self.path_pts,'')
            comb_0_a_naming=comb_0_a_naming.replace('.txt','')
            comb_1_b_naming = comb_1.replace(self.path_pts,'')
            comb_1_b_naming=comb_1_b_naming.replace('.txt','')
            self.combination_name = './pt_diff_'+str(comb_0_a_naming)+'_'+str(comb_1_b_naming)+'.txt'
            combination_name_no_txt = self.combination_name.replace('.txt','')

            if total_frame > 1000:
                job_excel = writexlsx(combination_name_no_txt)
                job_excel.add_info(comb_0,comb_1,pt_a_starting_frame,pt_b_starting_frame,common_starting_pt,int(len(list_pt_a)),int(len(list_pt_b)),total_frame) # <--

                list_pt_diff = self.pt_diff(list_pt_a,list_pt_b,common_diff,flag_starting_frame,total_frame)

                job_excel.add_data(list_pt_a,list_pt_b)
                job_excel.add_distance(list_pt_diff[0])
                job_excel.add_plot_graph()

                job_img=commomStartingImg(self.switch_lineMethod,common_starting_pt,comb_0,comb_1)
                insert_img=job_img.cam(list_pt_diff[1],list_pt_diff[2])   #<----need to add  0 or 1
                job_excel.add_img(insert_img,str('./graph/'+str(comb_0_a_naming)+'.png'),str('./graph/'+str(comb_1_b_naming)+'.png'))
                os.remove(insert_img)
                
    def pt_diff(self,l_p_a,l_p_b,c_d,f_s_f,t_f): #-->(list_pt_a,list_pt_b,common_starting_pt,flag_starting_frame,total_frame)
        list_diff = []
        if f_s_f == 0 or f_s_f == 2 :
            list_base = l_p_a
            list_other = l_p_b

            first_pt_a =  l_p_a[0]
            first_pt_b =  l_p_b[c_d]

        if f_s_f == 1:
            list_base = l_p_b
            list_other = l_p_a

            first_pt_a =  l_p_a[c_d]
            first_pt_b =  l_p_b[0]

        list_diff_result = []
        for i in range (0,t_f-1):
            i_other = i+c_d
            diff_x = float(list_base[i][0])-float(list_other[i_other][0])
            diff_y = float(list_base[i][1])-float(list_other[i_other][1])

            # diff_x = math.sqrt(diff_x*diff_x) # refreshs diff_x
            # diff_y = math.sqrt(diff_y*diff_y) # refreshs diff_y
            diff_result = math.sqrt((diff_x*diff_x)+(diff_y*diff_y))

            list_diff.append([diff_x,diff_y])
            list_diff_result.append(diff_result)

        
        with open(self.combination_name, "w", encoding="utf-8") as f3:
            for term_num in range(0,int(len(list_diff))):

                # ↓ diff in x , diff in y
                f3_line = str(list_diff[term_num][0])+"\t"+str(list_diff[term_num][1])+"\n"

                f3.write(f3_line)
            
        f3.close()
        list_return = [list_diff_result,first_pt_a,first_pt_b]
        return list_return


            

    def globalFrame(self):
        with open(self.path_global_frame) as f1:
            for line in f1.readlines():
                # print(line)    # testing
                # number = [int(temp)for temp in line.split() if temp.isdigit()]
                coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                frame = re.findall('\d+',line)
                global_info = (float(coor[0]),float(coor[1]),int(frame[len(frame)-1]))
                self.list_global.append(global_info)
        
        # print(self.list_global)
        # print(len(self.list_global))
        # print(self.list_global[0])

        # print(self.list_global[0][0])
        # print(type(self.list_global[0][0]))

        # print(self.list_global[0][1])
        # print(type(self.list_global[0][1]))
        # print(float(self.list_global[0][1]))   
        # print(type(float(self.list_global[0][1]))) # need to change type

        # print(self.list_global[0][2])
        # print(type(self.list_global[0][2]))

    def getPtsComb(self):
        self.list_pts = glob.glob(self.path_glob_pts)
        self.list_pts_comb = combinations(self.list_pts,2)
        # for i in self.list_pts_comb:
        #     print(i)
        # print('stop')

    def getPtsCombV2(self):
        self.list_pts_base = glob.glob(self.path_glob_pts_base)

        for comb2 in glob.glob(self.path_glob_pts_base):
            # job of base and region 1
            for comb1_2 in glob.glob(self.path_glob_pts_1):
                self.list_pts_comb1_2.insert([comb2,comb1_2])

            # job of base and region 3
            for comb3_2 in glob.glob(self.path_glob_pts_3):
                self.list_pts_comb3_2.insert([comb2,comb3_2])

            # job of base and region 4
            for comb4_2 in glob.glob(self.path_glob_pts_4):
                self.list_pts_comb4_2.insert([comb2,comb4_2])


#=============================
# init --> globalFrame (self.list_global) --> getPtsComb (self.list_pts_comb) --> run

class writexlsx:

    def __init__(self,excel_name) -> None:
        wb = Workbook()
        ws1 = wb.create_sheet("info")
        ws1.sheet_properties.tabColor = "1072BA"

        # insert date
        ws1['A1'] = 'Date:'
        datetime_dt = datetime.today()# 獲得當地時間
        datetime_str = datetime_dt.strftime("%Y/%m/%d %H:%M:%S")
        ws1['B1'] = datetime_str
        
        self.name = './'+str(excel_name)+'.xlsx'
        wb.save(self.name)


    def add_info(self,pt_a_path,pt_b_path,a_starting_frame,b_starting_frame,common_starting_frame,a_total,b_total,common_total):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['info']
        ws['A2']='data_set_a:'
        ws['A3']='data_set_b:'
        ws['A4']='pt_a_starting_frame:'
        ws['A5']='pt_b_starting_frame:'
        ws['A6']='common_starting_frame:'
        ws['A7']='pt_a_total_frame:'
        ws['A8']='pt_b_total_frame:'
        ws['A9']='common_total_frame:'

        ws['B2']=pt_a_path
        ws['B3']=pt_b_path
        ws['B4']=a_starting_frame
        ws['B5']=b_starting_frame
        ws['B6']=common_starting_frame
        ws['B7']=a_total
        ws['B8']=b_total
        ws['B9']=common_total

        self.a_start= int(a_starting_frame)
        self.b_start= int(b_starting_frame)
        self.common_frame = int(common_starting_frame)
        self.common_frame_stop = int(common_starting_frame)+int(common_total)
        wb.save(self.name)

    def add_data(self,list_a,list_b): 
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb.create_sheet('data')
        
        ws['A1'] = 'global_frame'
        ws['B1'] = 'pt_a_x'
        ws['C1'] = 'pt_b_x'
        ws['D1'] = 'pt_a_y'
        ws['E1'] = 'pt_b_y'
        ws['F1'] = 'distance'
         

        # frame number
        for k in range (0,3000):
            str_frame = str(k)
            pos_col_a = 'A'+str(k+2)
            ws[pos_col_a] = str_frame

        # add feature point data x and y
        for i in range (0,len(list_a)):
            pos_col_b = 'B'+str(i+2+self.a_start)
            pos_col_d = 'D'+str(i+2+self.a_start)
            ws[pos_col_b] = list_a[i][0]
            ws[pos_col_d] = list_a[i][1]

        # add feature point data x and y
        for j in range (0,len(list_b)):
            pos_col_c = 'C'+str(j+2+self.b_start)
            pos_col_e = 'E'+str(j+2+self.b_start)
            ws[pos_col_c] = list_b[j][0]
            ws[pos_col_e] = list_b[j][1]

        # for rows in list_a:
        #     ws.append(rows)       
        wb.save(self.name)


    def add_distance(self,list_distance):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        
        for i in range (0,len(list_distance)):
            pos_col_f = 'F'+str(i+2+self.common_frame)
            ws[pos_col_f] = list_distance[i]
        wb.save(self.name)


    def add_plot_graph(self): # plot graph 
        file_home = self.name
        wb = load_workbook(filename=file_home)
        ws = wb['data']
        chart = ScatterChart()
        chart.title = "Change of Distance"
        chart.style = 5
        chart.x_axis.title = 'Frame'
        chart.y_axis.title = 'Distance (Pixel)'

        xvalues = Reference(ws, min_col=1, min_row=2, max_row=int(self.common_frame_stop-1+2)) # range of x value

        values = Reference(ws, min_col=6, min_row=1, max_row=int(self.common_frame_stop-1+2)) # range of y value
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

        # 第一條散點
        s1 = chart.series[0]
        # 散點標記型別  'auto', 'dash', 'triangle', 'square', 'picture', 'circle', 'dot', 'plus', 'star', 'diamond', 'x'
        # s1.marker.symbol = "circle"
        # s1.marker.graphicalProperties.solidFill = "0000FF"  # Marker filling 設定標記填充的顏色
        # s1.marker.graphicalProperties.line.solidFill = "0000FF"  # Marker outline 標記輪廓的顏色
        # s1.graphicalProperties.line.noFill = True  # 關閉連線填充
        # s1.graphicalProperties

        # s1.marker.symbol = "circle"
        s1.graphicalProperties.solidFill = "FF0000"
        s1.marker.graphicalProperties.line.solidFill = "FF0000"
        s1.graphicalProperties.dashStyle = "dash"
        s1.smooth = True
        # s1.graphicalProperties.line.width = 1000  # width in EMUs

        ws.add_chart(chart, "A10")
        wb.save(self.name)

    def add_img(self,img,img2,img3):
        file_home = self.name
        wb = load_workbook(filename=file_home)
        # ws = wb['image']
        ws = wb.create_sheet('image')
        # image size --> (577*648)
        ws.column_dimensions['A'].width = 557*0.14
        ws.row_dimensions[1].height = 648 * 0.78
        img_toadd = Image(img)
        ws.add_image(img_toadd, 'A1')

        # image size --> (640*640)
        ws.column_dimensions['B'].width = 640*0.14
        ws.row_dimensions[1].height = 648 * 0.78
        img_toadd2 = Image(img2)
        ws.add_image(img_toadd2, 'B1')

        # image size --> (640*640)
        ws.column_dimensions['C'].width = 640*0.14
        ws.row_dimensions[1].height = 648 * 0.78
        img_toadd3 = Image(img3)
        ws.add_image(img_toadd3, 'C1')

        wb.save(self.name)

#========================================================================

class commomStartingImg:
    def __init__(self,switch,common_starting_pt,path_pt_a,path_pt_b) -> None:  # <---- comb0 and comb1 = path of pt a and pt b
        
        self.switch = int(switch)
        self.green_color = (0,255,0)
        self.scaleLine_width = 1
        self.text_x = 'X(+ve)'+'__'+str(common_starting_pt)
        self.text_y = 'Y(+ve)'

        if self.switch == 0:
            a_name = path_pt_a.replace('./after_first_filter/feature_point_0/','')
            self.a_name = a_name.replace('.txt','')
            b_name = path_pt_b.replace('./after_first_filter/feature_point_0/','')
            self.b_name = b_name.replace('.txt','')
        if self.switch == 1:
            a_name = path_pt_a.replace('./after_first_filter/feature_point_1/','')
            self.a_name = a_name.replace('.txt','')
            b_name = path_pt_b.replace('./after_first_filter/feature_point_1/','')
            self.b_name = b_name.replace('.txt','')

        self.output = './pt_diff_img_'+str(self.a_name)+'_'+str(self.b_name)+'.bmp'
        
        self.c_s_p = common_starting_pt
        self.path_a = path_pt_a
        self.path_b = path_pt_b

    def cam(self,coor_a,coor_b):

        if self.switch == 0:
            glob_bmp = './temp_img_0/*_0.bmp'
            eliminate2 = './temp_img_0/'
            eliminate3 = 'C_0.bmp'
            folder_temp_img = './temp_img_0/'
            folder_temp_img2 = 'C_0.bmp'
        if self.switch == 1:
            glob_bmp = './temp_img_1/*_1.bmp'
            eliminate2 = './temp_img_1/'
            eliminate3 = 'C_1.bmp'
            folder_temp_img = './temp_img_1/'
            folder_temp_img2 = 'C_1.bmp'

        list_img=[]
        for img in glob.glob(glob_bmp):
            img = str(img).replace('\\','/')
            img= img.replace(eliminate2,'')
            img=img.replace(eliminate3,'')
            list_img.append(img)

        list_img.sort()
        path_common_img = list_img[self.c_s_p]
        # C_0.bmp
        path_common_img = folder_temp_img+path_common_img+folder_temp_img2


        self.img = cv2.imread(path_common_img)
        cv2.line(self.img,(50,50),(50,100),self.green_color,self.scaleLine_width)
        cv2.line(self.img,(50,50),(100,50),self.green_color,self.scaleLine_width)
        cv2.line(self.img,(100,50),(90,40),self.green_color,self.scaleLine_width)
        cv2.line(self.img,(100,50),(90,60),self.green_color,self.scaleLine_width)
        cv2.line(self.img,(50,100),(40,90),self.green_color,self.scaleLine_width)
        cv2.line(self.img,(50,100),(60,90),self.green_color,self.scaleLine_width)
        cv2.putText(self.img, 'Y(+ve)', (20, 115), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
        cv2.putText(self.img, 'X(+ve)', (105, 55), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)

        cv2.circle(self.img,(int(float(coor_a[0])), int(float(coor_a[1]))),2,self.green_color, -1)
        cv2.putText(self.img, self.a_name, (int(float(coor_a[0]))+5, int(float(coor_a[1]))+5), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
        
        cv2.circle(self.img,(int(float(coor_b[0])), int(float(coor_b[1]))),2,self.green_color, -1)
        cv2.putText(self.img, self.b_name, (int(float(coor_b[0]))+5, int(float(coor_b[1]))+5), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
        
        cv2.imwrite(self.output, self.img)
        return self.output

    # def cam1(self,coor_a,coor_b):
    #     list_img2=[]
    #     for img in glob.glob('./temp_img_1/*_1.bmp'):
    #         img = str(img).replace('\\','/')
    #         img= img.replace('./temp_img_1/','')
    #         img=img.replace('C_1.bmp','')
    #         list_img2.append(img)

    #     list_img2.sort()
    #     path_common_img = list_img2[self.c_s_p]

    #     # C_1.bmp
    #     path_common_img = './temp_img_1/'+path_common_img+'C_1.bmp'

    #     self.img = cv2.imread(path_common_img)
    #     cv2.line(self.img,(50,50),(50,100),self.green_color,self.scaleLine_width)
    #     cv2.line(self.img,(50,50),(100,50),self.green_color,self.scaleLine_width)
    #     cv2.line(self.img,(100,50),(90,40),self.green_color,self.scaleLine_width)
    #     cv2.line(self.img,(100,50),(90,60),self.green_color,self.scaleLine_width)
    #     cv2.line(self.img,(50,100),(40,90),self.green_color,self.scaleLine_width)
    #     cv2.line(self.img,(50,100),(60,90),self.green_color,self.scaleLine_width)
    #     cv2.putText(self.img, 'Y(+ve)', (20, 115), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
    #     cv2.putText(self.img, 'X(+ve)', (105, 55), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
    #     cv2.circle(self.img,(int(float(coor_a[0])), int(float(coor_a[1]))),2,self.green_color, -1)
    #     cv2.putText(self.img, self.a_name, (int(float(coor_a[0]))+5, int(float(coor_a[1]))+5), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
    #     cv2.circle(self.img,(int(float(coor_b[0])), int(float(coor_b[1]))),2,self.green_color, -1)
    #     cv2.putText(self.img, self.b_name, (int(float(coor_b[0]))+5, int(float(coor_a[1]))+5), cv2.FONT_HERSHEY_PLAIN,1, self.green_color, 1, cv2.LINE_AA)
    #     cv2.imwrite(self.output, self.img)
    #     return self.output

#========================================================================
# job_a = lineMethod(1)
# job_a.globalFrame()
# job_a.getPtsComb()
# job_a.run()