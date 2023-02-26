import os
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def create_sheet(path4save):
    wb=load_workbook(path4save)
    wb.create_sheet('Trending',1)

    wb.save(path4save)

    print('Done create_sheet')

def do_trending(path_of_data):
    wb=load_workbook(path_of_data)
    ws_data = wb['DATA']
    ws_trending = wb['Trending']

    # getting data from 'DATA'
    trending = 'nan'
    bg_color = '#FFFFFF'
    for col in range (1,int(ws_data.max_column)+1):
        for row in range (2,int(ws_data.max_row)): # row does not +1 , cuz the last data does not has next data to compair
            # copying the case name
            ws_trending.cell(row=1,column=col,value=str(ws_data.cell(row=1,column=col).value))
            # do trending
            if ws_data.cell(row=row,column=col).value != None and ws_data.cell(row=row+1,column=col).value != None:
                n_th_data = float(ws_data.cell(row=row,column=col).value)
                nPlus1_th_data = float(ws_data.cell(row=row+1,column=col).value)
                compare_result = nPlus1_th_data-n_th_data
                if compare_result>0:
                    trending = 'up'
                    bg_color = '35FC03'
                if compare_result < 0:
                    trending='down'
                    bg_color = 'FC2C03'
                if compare_result ==0:
                    trending = 'nonChange'
                    bg_color = 'FCBA03'
                
                ws_trending.cell(row=row,column=col,value=str(trending))
                ws_trending.cell(row=row,column=col).fill = PatternFill(patternType='solid',fgColor=bg_color)
    
    wb.save(path_of_data)


    print('Done do_trending')


create_sheet('/Users/user/Documents/GitHub/backup/005_0025_001/area_method/DATA_func_test_bkup/DATA_0/DATA_0.xlsx')
do_trending('/Users/user/Documents/GitHub/backup/005_0025_001/area_method/DATA_func_test_bkup/DATA_0/DATA_0.xlsx')
