import glob
import logging
import re
from itertools import combinations
import math
import cv2
import os


from openpyxl.drawing.image import Image

from  openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)


class regMethod:
    def __init__(self,cam_num) -> None:
        
        if cam_num == 0:
            self.path_folder_coord = './after_first_filter/feature_point_0'
            self.path_folder_base = './after_first_filter/feature_point_0/region_2'
            self.path_global_frame = './feature_point_0/global_frame_number_0.txt'

            #create excel folder for cam_0
            if os.path.exists('./excel_cam_0'):
                logging.info('./excel_cam_0 has been created')
            else:
                os.system('mkdir ./excel_cam_0')
            self.path_excel = './excel_cam_0'

        if cam_num == 1:
            self.path_folder_coord = './after_first_filter/feature_point_1'
            self.path_folder_base = './after_first_filter/feature_point_1/region_2'
            self.path_global_frame = './feature_point_1/global_frame_number_1.txt'

            #create excel folder for cam_1
            if os.path.exists('./excel_cam_1'):
                logging.info('./excel_cam_1 has been created')
            else:
                os.system('mkdir ./excel_cam_1')
            self.path_excel = './excel_cam_1'
    
        pass

    def catchR2(self):

        golb_path_coord = self.path_folder_base+'/*.txt'

        self.list_global=[]
        self.getGlobalFrame()

        for txt in glob.glob(golb_path_coord):
            path_txt = txt
            txt_name = str(txt).split('/')[-1]
            name = txt_name.replace('.txt','')
            writexlsx(path_txt, self.list_global,name,self.path_excel, self.path_folder_coord).insertR2Data()


    def getGlobalFrame(self):
        with open(self.path_global_frame) as f1:
            for line in f1.readlines():
                # print(line)    # testing
                # number = [int(temp)for temp in line.split() if temp.isdigit()]
                coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                frame = re.findall('\d+',line)
                global_info = (float(coor[0]),float(coor[1]),int(frame[len(frame)-1]))
                self.list_global.append(global_info)


class writexlsx:

    def __init__(self,txt_path,global_list,excel_name,path_excel_folder, path_R_folder) -> None:

        self.this_path_txt = txt_path
        self.this_list_global = global_list
        self.this_path_R_folder = path_R_folder

        wb = Workbook()
        ws1 = wb.create_sheet("info")
        ws1.sheet_properties.tabColor = "1072BA"

        # insert date
        ws1['A1'] = 'Date:'
        datetime_dt = datetime.today()# 獲得當地時間
        datetime_str = datetime_dt.strftime("%Y/%m/%d %H:%M:%S")
        ws1['B1'] = datetime_str
        
        self.name = path_excel_folder+'/'+str(excel_name)+'.xlsx'
        wb.save(self.name)




    def insertR2Data(self):
        wb = load_workbook(self.name)
        ws_data = wb.create_sheet("BASE_vs_R1")
        ws_R3 = wb.create_sheet("BASE_vs_R3")
        ws_R4 = wb.create_sheet("BASE_vs_R4")

        #BASE
        #content
        ws_data.cell(row=1,column=1,value='BASE') # (R1)
        ws_data.cell(row=2,column=1,value='X') # (R1)
        ws_data.cell(row=2,column=2,value='Y') # (R1)
        ws_R3.cell(row=1,column=1,value='BASE') # (R3)
        ws_R3.cell(row=2,column=1,value='X') # (R3)
        ws_R3.cell(row=2,column=2,value='Y') # (R3)
        ws_R4.cell(row=1,column=1,value='BASE') # (R4)
        ws_R4.cell(row=2,column=1,value='X') # (R4)
        ws_R4.cell(row=2,column=2,value='Y') # (R4)

        #get BASE coord to list
        list_base = []
        with open(self.this_path_txt) as f1:
            for line in f1.readlines():
                # print(line)    # testing
                # number = [int(temp)for temp in line.split() if temp.isdigit()]
                coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                list_base.append(coor) # it contains whole coor of point BASE

        # get BASE global frame
        for i in range (len(self.this_list_global)):
            if float(list_base[0][0]) == float(self.this_list_global[i][0]) and float(list_base[0][1]) == float(self.this_list_global[i][1]):
                logging.info('x,y,frame_base: ',self.this_list_global[i]) # testing
                self.base_starting_frame = int(self.this_list_global[i][2]) # getting the starting frame number of BASE
                logging.info('base_starting_frame:',self.base_starting_frame) 
                break

        #insert BASE data
        for base in range (0,len(list_base)):
            ws_data.cell(row=(2+1+self.base_starting_frame+base),column=1,value=list_base[base][0]) # BASE X (R1)
            ws_data.cell(row=(2+1+self.base_starting_frame+base),column=2,value=list_base[base][1]) # BASE Y (R1)
            ws_R3.cell(row=(2+1+self.base_starting_frame+base),column=1,value=list_base[base][0]) # BASE X (R3)
            ws_R3.cell(row=(2+1+self.base_starting_frame+base),column=2,value=list_base[base][1]) # BASE Y (R3)
            ws_R4.cell(row=(2+1+self.base_starting_frame+base),column=1,value=list_base[base][0]) # BASE X (R4)
            ws_R4.cell(row=(2+1+self.base_starting_frame+base),column=2,value=list_base[base][1]) # BASE Y (R4)

        # R1
        path_R1 = self.this_path_R_folder+'/region_1/*.txt'
        R1_count = 0
        for R1_feature_pt in glob.glob(path_R1):
            name_R1_feature_pt = str(R1_feature_pt).split('/')[-1]
            name_R1_feature_pt = name_R1_feature_pt.replace('.txt','')
            list_fea_pt = []
            with open(R1_feature_pt) as f1:
                for line in f1.readlines():
                    # print(line)    # testing
                    # number = [int(temp)for temp in line.split() if temp.isdigit()]
                    coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                    list_fea_pt.append(coor) 
            
            # get R1 global frame
            R1_starting_frame = 9999
            for i in range (len(self.this_list_global)):
                if float(list_fea_pt[0][0]) == float(self.this_list_global[i][0]) and float(list_fea_pt[0][1]) == float(self.this_list_global[i][1]):
                    logging.info('x,y,frame_R1_'+name_R1_feature_pt+': ',self.this_list_global[i]) # testing
                    R1_starting_frame = int(self.this_list_global[i][2]) # getting the starting frame number of BASE
                    logging.info('base_starting_frame:',R1_starting_frame) 
                    break

            ws_data.cell(row=1,column=(3+(R1_count*2)),value=name_R1_feature_pt)
            ws_data.cell(row=2,column=(3+(R1_count*2)),value='X')
            ws_data.cell(row=2,column=(4+(R1_count*2)),value='Y')
            # insert R1 data pt
            for coord_term in range (0,len(list_fea_pt)):
                ws_data.cell(row=(2+1+R1_starting_frame+coord_term),column=(3+(R1_count*2)),value=list_fea_pt[coord_term][0]) # X
                ws_data.cell(row=(2+1+R1_starting_frame+coord_term),column=(4+(R1_count*2)),value=list_fea_pt[coord_term][1]) # Y

            R1_count+=1
        
        # R3
        path_R3 = self.this_path_R_folder+'/region_3/*.txt'
        R3_count = 0
        for R3_feature_pt in glob.glob(path_R3):
            name_R3_feature_pt = str(R3_feature_pt).split('/')[-1]
            name_R3_feature_pt = name_R3_feature_pt.replace('.txt','')
            list_fea_pt_r3 = []
            with open(R3_feature_pt) as f1:
                for line in f1.readlines():
                    # print(line)    # testing
                    # number = [int(temp)for temp in line.split() if temp.isdigit()]
                    coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                    list_fea_pt_r3.append(coor) 
            
            # get R1 global frame
            R3_starting_frame = 9999
            for i in range (len(self.this_list_global)):
                if float(list_fea_pt_r3[0][0]) == float(self.this_list_global[i][0]) and float(list_fea_pt_r3[0][1]) == float(self.this_list_global[i][1]):
                    logging.info('x,y,frame_R1_'+name_R3_feature_pt+': ',self.this_list_global[i]) # testing
                    R3_starting_frame = int(self.this_list_global[i][2]) # getting the starting frame number of BASE
                    logging.info('base_starting_frame:',R3_starting_frame) 
                    break

            ws_R3.cell(row=1,column=(3+(R3_count*2)),value=name_R3_feature_pt)
            ws_R3.cell(row=2,column=(3+(R3_count*2)),value='X')
            ws_R3.cell(row=2,column=(4+(R3_count*2)),value='Y')
            # insert R1 data pt
            for coord_term in range (0,len(list_fea_pt_r3)):
                ws_R3.cell(row=(2+1+R3_starting_frame+coord_term),column=(3+(R3_count*2)),value=list_fea_pt_r3[coord_term][0]) # X
                ws_R3.cell(row=(2+1+R3_starting_frame+coord_term),column=(4+(R3_count*2)),value=list_fea_pt_r3[coord_term][1]) # Y

            R3_count+=1
        
        # R4
        path_R4 = self.this_path_R_folder+'/region_4/*.txt'
        R4_count = 0
        for R4_feature_pt in glob.glob(path_R4):
            name_R4_feature_pt = str(R4_feature_pt).split('/')[-1]
            name_R4_feature_pt = name_R4_feature_pt.replace('.txt','')
            list_fea_pt_r4 = []
            with open(R4_feature_pt) as f1:
                for line in f1.readlines():
                    # print(line)    # testing
                    # number = [int(temp)for temp in line.split() if temp.isdigit()]
                    coor = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                    list_fea_pt_r4.append(coor) 
            
            # get R1 global frame
            R4_starting_frame = 9999
            for i in range (len(self.this_list_global)):
                if float(list_fea_pt_r4[0][0]) == float(self.this_list_global[i][0]) and float(list_fea_pt_r4[0][1]) == float(self.this_list_global[i][1]):
                    logging.info('x,y,frame_R1_'+name_R4_feature_pt+': ',self.this_list_global[i]) # testing
                    R4_starting_frame = int(self.this_list_global[i][2]) # getting the starting frame number of BASE
                    logging.info('base_starting_frame:',R4_starting_frame) 
                    break

            ws_R4.cell(row=1,column=(3+(R4_count*2)),value=name_R4_feature_pt)
            ws_R4.cell(row=2,column=(3+(R4_count*2)),value='X')
            ws_R4.cell(row=2,column=(4+(R4_count*2)),value='Y')
            # insert R1 data pt
            for coord_term in range (0,len(list_fea_pt_r4)):
                ws_R4.cell(row=(2+1+R4_starting_frame+coord_term),column=(3+(R4_count*2)),value=list_fea_pt_r4[coord_term][0]) # X
                ws_R4.cell(row=(2+1+R4_starting_frame+coord_term),column=(4+(R4_count*2)),value=list_fea_pt_r4[coord_term][1]) # Y

            R4_count+=1

        wb.save(self.name)




#==============================
job=regMethod(1)
job.catchR2()

