import glob
import logging
import re
from itertools import combinations
import math
import shutil 
import os


from openpyxl.drawing.image import Image

from  openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class regionIntegration:
    
    def __init__(self) -> None:
        #create folder region_134
        if os.path.exists('./excel_cam_0/region_134')==False:
            os.system('mkdir ./excel_cam_0/region_134')
        if os.path.exists('./excel_cam_1/region_134')==False:
            os.system('mkdir ./excel_cam_1/region_134')
        
        for cam in (['./excel_cam_0','./excel_cam_1']):
        # for cam in (['./excel_cam_1']):
            for region in (['region_1','region_3','region_4']):
                if os.path.exists(cam+'/'+region) == True:
                    src_files = os.listdir(cam+'/'+region)
                    for file_name in src_files:
                        full_file_name = os.path.join(cam+'/'+region, file_name)
                        if os.path.isfile(full_file_name):
                            shutil.copy(full_file_name, cam+'/region_134')


class writexlsx:

    def __init__(self) -> None:
        # # insert date
        # ws1['A1'] = 'Date:'
        # datetime_dt = datetime.today()# 獲得當地時間
        # datetime_str = datetime_dt.strftime("%Y/%m/%d %H:%M:%S")
        # ws1['B1'] = datetime_str
        
        # self.name = path_excel_folder+'/'+str(excel_name)+'.xlsx'
        # wb.save(self.name)
        for cam in (['./excel_cam_0','./excel_cam_1']):
        # for cam in (['./excel_cam_1']):
            for region in (['region_1','region_3','region_4','region_134']):
                if os.path.exists(cam+'/'+region) == True:
                    ls_base = [] 
                    for base_pt in glob.glob(cam+'/'+region+'/*.txt'):
                        base_pt = base_pt.replace(cam+'/'+region+'/','')
                        base_pt = base_pt.split('_')[0]
                        if base_pt not in ls_base:
                            ls_base.append(base_pt)

                    for pt in ls_base:
                        wb = Workbook()
                        ws1 = wb.create_sheet("Data")
                        ws1.sheet_properties.tabColor = "1072BA"
                        case_col = 1
                        for case in glob.glob(cam+'/'+region+'/*.txt'):
                            name_split_case = case.replace(cam+'/'+region+'/','')
                            name_split_case_0 = name_split_case.split('_')[0]
                            if str(name_split_case_0) == str(pt):
                                case_name = str(case).replace(cam+'/'+region+'/','')
                                case_name = case_name.replace('.txt','')
                                start_frame = case_name.split('_')[6]
                                ws1.cell(row=1,column=case_col,value=case_name)
                                with open(case) as f1:
                                    line_num = 1
                                    for line in f1.readlines():
                                        # disp = re.findall("\d+\.\d+", line)    # format:string    [convect from string with dot to float]
                                        ws1.cell(row=int(start_frame)+line_num,column=case_col,value=float(line))
                                        line_num+=1
                                case_col+=1
                    
                        cam_name = cam.replace('./','')
                        wb.save(cam+'/'+region+'/'+cam_name+'_'+region+'_'+pt+'.xlsx')

#==============================
