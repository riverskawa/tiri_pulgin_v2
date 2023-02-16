import os
import glob
from openpyxl import load_workbook
from openpyxl import Workbook
from multiprocessing.dummy import Pool as ThreadPool
import shutil

def create_xlsx(path4save):
    
    wb = Workbook()
    wb.create_sheet('DATA',0)
    wb.create_sheet('GRAPH',1)

    wb.save(path4save)

    print('Done create_xlsx')


def data_fill_in(path_xlsx2edit,ls_data):
    wb=load_workbook(path_xlsx2edit)
    sheet_editing = wb['DATA']

    all_col = 1
    for term_case in ls_data:
        
        #obtain starting frame
        area_start_frame=int(term_case[0])
        case_name = str(term_case[1])
        sheet_editing.cell(row=1,column =all_col,value=case_name)
        for data2fill in range (0,int(len(term_case)-15)):
            sheet_editing.cell(row=int(data2fill+2+area_start_frame),column =all_col,value=term_case[data2fill+15])    # the actual area common starting frame is excel_row-2
        
        all_col+=1
    
    wb.save(path_xlsx2edit)
    
    # #

    # sheet_editing.cell(row=1,column =col2start,value=str(case))
    # for frame in range (0,len(ls_data)-14): # 14 = -1 -13(txt file data starts from #13) 
    #     sheet_editing.cell(row=int(row2start)+2+frame,column=col2start,value = ls_data[frame+13])
    # wb.save(path_xlsx2edit)
    
    # print('Done data_fill_in '+str(case))




def get_data_from_txt(path_data_folder):

    path4glob = path_data_folder+'/*.txt'
    ls_txt = glob.glob(path4glob)

    # txt_counter = 1 # setting it start at 1 (xlsx counter begins from 1 )

    path_xlsx_output = path_data_folder+'/'+str(path_data_folder.split('/')[-1])+'.xlsx'

    create_xlsx(path_xlsx_output)

    ls_allTxt = []
    for txt in ls_txt:
        
        ls_line=[]
        f = open(txt, 'r')
        for line in f.readlines():
            line=line.replace('\n','')
            ls_line.append(line)
        f.close() 

        area_starting_frame=int(str(ls_line[10]).split(':')[1])
        case_name = txt.split('/')[-1]
        case_name = str(case_name).replace('.txt','')

        ls_line.insert(0,area_starting_frame)
        ls_line.insert(1,case_name)
        # !!! ls_line[int(area_starting_frame,str(case_name),.......)] <---- structure

        ls_allTxt.append(ls_line)


        # do data fill in to the xlsx
    data_fill_in(path_xlsx_output,ls_allTxt)
    print('Done get_data_from_txt '+str(path_data_folder))


def splitting_data(path_data_folder):
    path4glob = path_data_folder+'/*.txt'
    ls_txt = glob.glob(path4glob)

    len_folder=0
    if len(ls_txt)>=1000:
        len_folder = len(ls_txt)/1000
        if len_folder < 1:
            len_folder =1
        if len(ls_txt)%1000>0:
            len_folder=len_folder+1
    if len(ls_txt) < 1000:
        len_folder =1
    
    # create sub-folders
    ls_subfolder = []
    if len_folder > 1: 
        for folder_num in range (0,int(len_folder)):
            cmd_line = path_data_folder+'/DATA_'+str(folder_num)
            os.system('mkdir '+cmd_line)
            ls_subfolder.append(cmd_line)
    if len_folder == 1:
        os.system('mkdir '+str(path_data_folder)+'/DATA_0')
        ls_subfolder= [str(path_data_folder)+'/DATA_0']
    
 
    sub_folder_num2use = 0
    for ii in range (0,len(ls_txt)):

        shutil.move(ls_txt[ii],ls_subfolder[sub_folder_num2use])
        if ii != 0:
            if ii%999 == 0:
                sub_folder_num2use+=1
    
    return ls_subfolder # returning ls of full path
    
def main_data_present(path_folder):
    ls_subFolder=splitting_data(path_folder)
    p=ThreadPool(5)
    p.map(get_data_from_txt,ls_subFolder)

#=====independent test==============================================
main_data_present('/Users/user/Documents/GitHub/backup/005_0025_001/area_method/DATA')